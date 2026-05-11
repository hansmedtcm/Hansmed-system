"""
Brief #18 prep — Tier 1 TCM prescription guidance data template.
Public-domain classical knowledge (傷寒論, 金匱要略, 內經, 神農本草經 era texts).
User reviews + corrects + supplements with their textbook as reference.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = '/sessions/lucid-gallant-goldberg/mnt/Hansmed-system/reference/tcm-data-tier1.xlsx'
os.makedirs(os.path.dirname(OUT), exist_ok=True)

wb = Workbook()
wb.remove(wb.active)

H_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
H_FILL = PatternFill('solid', start_color='5C4A3C')
H_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
B_FONT = Font(name='Arial', size=10)
B_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)
ZEBRA = PatternFill('solid', start_color='F5F1EA')
THIN = Side(border_style='thin', color='D6CFC2')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def setup_sheet(ws, headers, widths):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = H_FONT
        c.fill = H_FILL
        c.alignment = H_ALIGN
        c.border = BORDER
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = 'A2'

def write_rows(ws, rows):
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font = B_FONT
            c.alignment = B_ALIGN
            c.border = BORDER
            if r_idx % 2 == 0:
                c.fill = ZEBRA
        ws.row_dimensions[r_idx].height = 60

# README
ws = wb.create_sheet('README')
ws.column_dimensions['A'].width = 110
readme = [
    'HansMed TCM Prescription Guidance — Tier 1 Data',
    '',
    'Brief #18 reference dataset. Public-domain classical TCM data only.',
    '',
    'WHAT THIS IS',
    'A starting reference for the doctor-side prescription guidance system. Five data tabs:',
    '  • Syndromes (證型) — 30 classical TCM syndromes with diagnostic criteria',
    '  • Formulas (方劑) — 50 classical formulas with composition + indications',
    '  • Herbs (藥材) — 80 commonly used herbs with properties + dose ranges',
    '  • Symptoms — controlled vocabulary so doctor input matches syndrome criteria',
    '  • RedFlags — absolute contraindications (pregnancy, herb-drug, halal concerns)',
    '',
    'COPYRIGHT / IP',
    'All data here is from public-domain classical sources (傷寒論, 金匱要略, 內經, 神農本草經 era).',
    'Compositions and classical indications of formulas like 桂枝湯, 小柴胡湯 are 1,800+ years old.',
    'Modern textbooks (e.g. 中醫治法與方劑 5th ed.) are copyrighted — use them as YOUR personal',
    'verification reference, but do NOT paste their specific phrasing into this spreadsheet.',
    '',
    'HOW TO USE',
    '1. Review each row. Cross-check with your textbook + clinical experience.',
    '2. Edit/correct cells as needed. Your house-style overrides classical defaults.',
    '3. Add Malaysian-context overlays: halal alternatives, locally-available herbs, T&CM Act notes.',
    '4. Add your own house formulas / proprietary blends in extra rows.',
    '5. When ready, this becomes the source of truth for the prescription guidance system.',
    '',
    'COLUMN CODES',
    '  ID — stable identifier; do not change once assigned (system uses it for foreign keys)',
    '  ZH — Chinese characters (traditional preferred for clinical TCM)',
    '  PY — Pinyin romanisation',
    '  Source — classical text the formula/syndrome traces to',
    '',
    'WHEN A CELL SAYS "VERIFY"',
    'Cells flagged where my training data felt less certain — cross-check those first.',
    '',
    'NEXT STEPS',
    '  • Brief #16 (email infra) ships first',
    '  • Brief #17 — doctor onboarding flow',
    '  • Brief #18 — this prescription guidance system, using this spreadsheet as input data',
    '',
    'Generated 6 May 2026 by Claude (Cowork mode) for HansMed Modern TCM.',
]
for i, line in enumerate(readme, 1):
    c = ws.cell(row=i, column=1, value=line)
    c.font = Font(name='Arial', size=11, bold=(i == 1))
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Syndromes
ws = wb.create_sheet('Syndromes')
setup_sheet(ws,
    ['ID','EN Name','ZH Name','PY','Category','Primary Symptoms (主證)','Secondary Symptoms (次證)','Tongue (舌象)','Pulse (脈象)','Treatment Principle (治法)','Default Formula(s)','Source','Notes'],
    [10, 30, 14, 22, 18, 50, 50, 25, 20, 35, 28, 16, 35])

syndromes = [
    ('SY01','Wind-Cold Tai Yang','太陽傷寒','Tai Yang Shang Han','Six-Channel — Cold','Aversion to cold; fever; no sweating; body aches; headache','Stiff neck; dyspnea; clear nasal discharge','Thin white coat','Floating tight','Release exterior; warm and dispel cold','麻黃湯 (Ma Huang Tang)','傷寒論',''),
    ('SY02','Wind Tai Yang','太陽中風','Tai Yang Zhong Feng','Six-Channel — Cold','Fever; aversion to wind; sweating; headache','Nasal congestion; mild cough','Thin white moist','Floating slow','Release exterior; harmonise ying-wei','桂枝湯 (Gui Zhi Tang)','傷寒論',''),
    ('SY03','Shao Yang','少陽證','Shao Yang Zheng','Six-Channel — Half','Alternating chills and fever; bitter mouth; chest fullness; nausea','Poor appetite; dizziness; dry throat','Thin white','Wiry','Harmonise shao yang','小柴胡湯 (Xiao Chai Hu Tang)','傷寒論',''),
    ('SY04','Yang Ming Excess','陽明腑實','Yang Ming Fu Shi','Six-Channel — Heat','High fever; constipation; abdominal fullness/tenderness; restlessness','Sweating; dry mouth; possible delirium','Yellow dry thick','Sinking forceful','Drain heat; purge fu organs','大承氣湯 (Da Cheng Qi Tang)','傷寒論','VERIFY contraindications carefully'),
    ('SY05','Tai Yin Deficiency Cold','太陰虛寒','Tai Yin Xu Han','Six-Channel — Cold','Abdominal fullness; vomiting; loose stool; poor appetite','Cold extremities; lassitude','Pale wet','Slow weak','Warm middle; dispel cold','理中湯 (Li Zhong Tang)','傷寒論',''),
    ('SY06','Shao Yin Cold Pattern','少陰寒化','Shao Yin Han Hua','Six-Channel — Cold','Aversion to cold; lying curled up; cold extremities; clear urine','Lassitude; weak pulse','Pale','Faint thready','Rescue yang; warm jueyin','四逆湯 (Si Ni Tang)','傷寒論','EMERGENCY pattern; refer if severe'),
    ('SY07','Liver Qi Stagnation','肝氣鬱結','Gan Qi Yu Jie','Zang-Fu — Liver','Hypochondriac distension; chest oppression; sighing; mood swings','Premenstrual tension; irritability','Thin white','Wiry','Soothe liver; rectify qi','柴胡疏肝散; 逍遙散','和劑局方',''),
    ('SY08','Liver Qi to Fire','肝鬱化火','Gan Yu Hua Huo','Zang-Fu — Liver','Irritability; bitter mouth; red eyes; headache; tinnitus','Insomnia; vivid dreams; constipation','Red with yellow coat','Wiry rapid','Clear liver; drain fire','龍膽瀉肝湯 (Long Dan Xie Gan Tang)','和劑局方',''),
    ('SY09','Liver Yang Rising','肝陽上亢','Gan Yang Shang Kang','Zang-Fu — Liver','Headache; dizziness; tinnitus; flushed face; irritability','Numbness in limbs; insomnia','Red','Wiry forceful','Calm liver; subdue yang','天麻鉤藤飲 (Tian Ma Gou Teng Yin)','雜病證治新義','VERIFY blood pressure measurements'),
    ('SY10','Liver Blood Deficiency','肝血虛','Gan Xue Xu','Zang-Fu — Liver','Pale dull complexion; blurred vision; tinnitus; numb limbs','Scant menstruation; insomnia','Pale','Thready','Nourish blood; tonify liver','四物湯 (Si Wu Tang)','和劑局方',''),
    ('SY11','Heart Blood Deficiency','心血虛','Xin Xue Xu','Zang-Fu — Heart','Palpitations; insomnia; poor memory; pale complexion','Anxiety; vivid dreams','Pale','Thready','Nourish blood; calm shen','歸脾湯 (Gui Pi Tang)','濟生方',''),
    ('SY12','Heart Yin Deficiency','心陰虛','Xin Yin Xu','Zang-Fu — Heart','Palpitations; insomnia; night sweats; five-centre heat','Mouth ulcers; dry throat; vivid dreams','Red with little coat','Thready rapid','Nourish yin; calm shen','天王補心丹 (Tian Wang Bu Xin Dan)','攝生秘剖',''),
    ('SY13','Spleen Qi Deficiency','脾氣虛','Pi Qi Xu','Zang-Fu — Spleen','Loose stool; bloating after eating; fatigue; pale complexion','Poor appetite; spontaneous sweat','Pale with teeth marks','Weak','Tonify spleen; boost qi','四君子湯; 六君子湯','和劑局方',''),
    ('SY14','Spleen Yang Deficiency','脾陽虛','Pi Yang Xu','Zang-Fu — Spleen','Cold abdominal pain (relieved by warmth/pressure); loose stool; cold extremities','Pale complexion; lassitude; clear urine','Pale wet','Sinking slow','Warm spleen; dispel cold','附子理中湯 (Fu Zi Li Zhong Tang)','和劑局方',''),
    ('SY15','Middle Qi Sinking','中氣下陷','Zhong Qi Xia Xian','Zang-Fu — Spleen','Visceral prolapse (uterine, rectal, gastric); chronic diarrhoea; SOB','Pale dull complexion; lassitude','Pale','Weak','Lift middle qi; tonify','補中益氣湯 (Bu Zhong Yi Qi Tang)','脾胃論',''),
    ('SY16','Stomach Yin Deficiency','胃陰虛','Wei Yin Xu','Zang-Fu — Stomach','Hunger without appetite; dry mouth; epigastric burning; constipation','Belching; nausea','Red dry no coat','Thready rapid','Nourish stomach yin','益胃湯 (Yi Wei Tang)','溫病條辨',''),
    ('SY17','Food Stagnation','食滯胃脘','Shi Zhi Wei Wan','Zang-Fu — Stomach','Epigastric distension; sour belching; foul-smelling stool; nausea','Anorexia; loose stool','Thick greasy','Slippery','Disperse food; rectify qi','保和丸 (Bao He Wan)','丹溪心法',''),
    ('SY18','Lung Qi Deficiency','肺氣虛','Fei Qi Xu','Zang-Fu — Lung','Weak voice; SOB; spontaneous sweat; susceptibility to colds','Pale complexion; fatigue','Pale','Weak','Tonify lung qi','玉屏風散 (Yu Ping Feng San)','丹溪心法',''),
    ('SY19','Lung Yin Deficiency','肺陰虛','Fei Yin Xu','Zang-Fu — Lung','Dry cough; little sticky sputum; hoarse voice; afternoon fever','Night sweats; dry throat','Red dry','Thready rapid','Nourish lung yin','沙參麥冬湯 (Sha Shen Mai Dong Tang)','溫病條辨',''),
    ('SY20','Wind-Cold Cough','風寒咳嗽','Feng Han Ke Sou','Zang-Fu — Lung','Cough with thin white sputum; nasal congestion; aversion to cold','Headache; body aches','Thin white','Floating tight','Disperse wind-cold; release lung qi','三拗湯; 止嗽散','和劑局方; 醫學心悟',''),
    ('SY21','Wind-Heat Lung','風熱犯肺','Feng Re Fan Fei','Zang-Fu — Lung','Cough with thick yellow sputum; sore throat; fever','Thirst; headache','Red tip thin yellow','Floating rapid','Disperse wind-heat; clear lung','桑菊飲; 銀翹散','溫病條辨',''),
    ('SY22','Kidney Yang Deficiency','腎陽虛','Shen Yang Xu','Zang-Fu — Kidney','Cold lower back / knees; impotence; clear copious urine; oedema','Lassitude; pale complexion','Pale wet','Sinking weak','Warm and tonify kidney yang','金匱腎氣丸 (Jin Gui Shen Qi Wan)','金匱要略',''),
    ('SY23','Kidney Yin Deficiency','腎陰虛','Shen Yin Xu','Zang-Fu — Kidney','Lower back / knee soreness; tinnitus; dizziness; afternoon flushing','Night sweats; nocturnal emission; insomnia','Red dry','Thready rapid','Nourish kidney yin','六味地黃丸 (Liu Wei Di Huang Wan)','小兒藥證直訣',''),
    ('SY24','Kidney Yin + Empty Heat','腎陰虛火旺','Shen Yin Xu Huo Wang','Zang-Fu — Kidney','As Kidney Yin Def + bone-steaming heat; spermatorrhoea; chronic sore throat','Severe night sweats; thirst','Red dry','Thready rapid','Nourish yin; clear empty heat','知柏地黃丸 (Zhi Bai Di Huang Wan)','醫宗金鑑',''),
    ('SY25','Qi + Blood Deficiency','氣血兩虛','Qi Xue Liang Xu','Eight Principles — Def','Pale complexion; fatigue; palpitations; insomnia; SOB','Dizziness; lassitude','Pale','Thready weak','Tonify qi and blood','八珍湯; 十全大補湯','正體類要',''),
    ('SY26','Phlegm-Damp','痰濕內蘊','Tan Shi Nei Yun','Pathological Factor','Productive cough with copious white sputum; fullness in chest; nausea','Dizziness; heavy body; bland taste','Thick greasy','Slippery','Dry damp; transform phlegm','二陳湯; 平胃散','和劑局方',''),
    ('SY27','Damp-Heat Lower Burner','濕熱下注','Shi Re Xia Zhu','Pathological Factor','Cloudy yellow urine; dysuria; vaginal/genital itch with discharge','Lower abdominal heat; constipation','Yellow greasy','Slippery rapid','Clear heat; drain damp','八正散; 龍膽瀉肝湯','和劑局方',''),
    ('SY28','Damp-Heat Middle Burner','濕熱中焦','Shi Re Zhong Jiao','Pathological Factor','Epigastric fullness; nausea; bitter mouth; jaundice possible','Heavy limbs; loose foul stool','Yellow greasy','Soft rapid','Clear heat; transform damp','三仁湯; 茵陳蒿湯','溫病條辨',''),
    ('SY29','Heart-Spleen Disharmony','心脾兩虛','Xin Pi Liang Xu','Zang-Fu — Combined','Palpitations; insomnia; poor memory; pale; loose stool; poor appetite','Vivid dreams; lassitude','Pale','Thready weak','Tonify heart-spleen; nourish blood','歸脾湯 (Gui Pi Tang)','濟生方',''),
    ('SY30','Liver-Spleen Disharmony','肝脾不和','Gan Pi Bu He','Zang-Fu — Combined','Hypochondriac distension; abdominal bloating; loose stool; mood swings','PMS; sighing; gas','Thin white','Wiry','Soothe liver; harmonise spleen','逍遙散; 痛瀉要方','和劑局方; 景岳全書',''),
]
write_rows(ws, syndromes)

# Formulas
ws = wb.create_sheet('Formulas')
setup_sheet(ws,
    ['ID','EN Name','ZH Name','PY','Category','Composition (herb : daily dose g)','Indications (Syndrome IDs)','Treatment Principle','Contraindications','Modifications','Source','Notes'],
    [10, 36, 14, 28, 18, 70, 16, 32, 32, 40, 14, 32])

formulas = [
    ('FM01','Cinnamon Twig Decoction','桂枝湯','Gui Zhi Tang','Release Exterior — Warm','桂枝 9; 白芍 9; 生薑 9; 大棗 4; 炙甘草 6','SY02','Release exterior; harmonise ying-wei','High fever without sweat; interior heat','Add 葛根 for stiff neck (桂枝加葛根湯)','傷寒論',''),
    ('FM02','Ephedra Decoction','麻黃湯','Ma Huang Tang','Release Exterior — Warm','麻黃 9; 桂枝 6; 杏仁 9; 炙甘草 3','SY01','Release exterior; warm and dispel cold','Sweating; weak constitution; hypertension','Add 石膏 for heat (麻杏石甘湯 redirect)','傷寒論','CAUTION cardiac patients'),
    ('FM03','Minor Bupleurum Decoction','小柴胡湯','Xiao Chai Hu Tang','Harmonise','柴胡 12; 黃芩 9; 半夏 9; 人參 9; 炙甘草 6; 生薑 9; 大棗 4','SY03','Harmonise shao yang','Pure exterior or pure interior pattern','Add 牡蠣 for chest fullness; Add 桂枝 if exterior remains','傷寒論',''),
    ('FM04','Major Order the Qi Decoction','大承氣湯','Da Cheng Qi Tang','Drain Downward','大黃 12; 厚朴 24; 枳實 12; 芒硝 9','SY04','Drain heat; purge fu organs','Pregnancy; weak elderly','Reduce dose for elderly (調胃承氣湯 alt)','傷寒論','EMERGENCY use only; REFER if uncertain'),
    ('FM05','Regulate the Middle Decoction','理中湯','Li Zhong Tang','Warm Interior','人參 9; 乾薑 9; 白朮 9; 炙甘草 6','SY05','Warm middle; tonify spleen-stomach yang','Yin deficiency; stomach heat','Add 附子 for severe cold (附子理中湯)','傷寒論',''),
    ('FM06','Aconite Regulate the Middle','附子理中湯','Fu Zi Li Zhong Tang','Warm Interior','附子 6-9; 人參 9; 乾薑 9; 白朮 9; 炙甘草 6','SY14','Warm spleen and kidney yang','Yin deficiency; pregnancy','—','和劑局方','附子 must be 先煎; toxicity if undercooked'),
    ('FM07','Frigid Extremities Decoction','四逆湯','Si Ni Tang','Rescue Yang','附子 9-15; 乾薑 9; 炙甘草 6','SY06','Rescue yang from devastation','Heat collapse (true heat false cold)','—','傷寒論','EMERGENCY; 附子 toxicity caution'),
    ('FM08','Bupleurum Liver-Coursing Powder','柴胡疏肝散','Chai Hu Shu Gan San','Regulate Qi','柴胡 6; 陳皮 6; 川芎 4.5; 香附 4.5; 枳殼 4.5; 白芍 4.5; 炙甘草 1.5','SY07','Soothe liver; rectify qi','Liver fire blazing','Add 香附 + 鬱金 for severe stagnation','景岳全書',''),
    ('FM09','Free and Easy Wanderer Powder','逍遙散','Xiao Yao San','Regulate Qi + Tonify','柴胡 9; 當歸 9; 白芍 9; 白朮 9; 茯苓 9; 炙甘草 4.5; 薄荷 3; 生薑 3','SY07; SY30','Soothe liver; strengthen spleen; nourish blood','Pure liver fire','加味逍遙散 adds 牡丹皮 + 梔子 for fire','和劑局方',''),
    ('FM10','Augmented Free and Easy','加味逍遙散','Jia Wei Xiao Yao San','Regulate Qi + Clear Heat','逍遙散 + 牡丹皮 6; 梔子 6','SY08 (mild); PMS','Soothe liver; clear heat; nourish blood','Severe yin deficiency','—','女科撮要',''),
    ('FM11','Gentian Drain the Liver','龍膽瀉肝湯','Long Dan Xie Gan Tang','Clear Heat','龍膽草 6; 黃芩 9; 梔子 9; 澤瀉 12; 木通 6; 車前子 9; 當歸 6; 生地 9; 柴胡 6; 甘草 3','SY08; SY27','Clear liver fire; drain damp-heat','Spleen-stomach deficiency; pregnancy','—','和劑局方','木通 use 川木通 (NOT 關木通 — nephrotoxic)'),
    ('FM12','Gastrodia Uncaria Drink','天麻鉤藤飲','Tian Ma Gou Teng Yin','Calm Wind','天麻 9; 鉤藤 12; 石決明 18; 山梔 9; 黃芩 9; 川牛膝 12; 杜仲 9; 益母草 9; 桑寄生 9; 夜交藤 9; 茯神 9','SY09','Calm liver; subdue yang; nourish kidney','Pure deficiency without yang rising','—','雜病證治新義','VERIFY BP'),
    ('FM13','Four Substances Decoction','四物湯','Si Wu Tang','Tonify Blood','當歸 9; 川芎 6; 白芍 9; 熟地 12','SY10; menstrual','Nourish blood; regulate menses','Heavy bleeding; severe damp','+四君子 = 八珍湯','和劑局方',''),
    ('FM14','Eight Treasures Decoction','八珍湯','Ba Zhen Tang','Tonify Qi+Blood','人參 9; 白朮 9; 茯苓 9; 炙甘草 6; 當歸 9; 川芎 6; 白芍 9; 熟地 12','SY25','Tonify qi and blood','Acute exterior pattern','+黃耆+肉桂 = 十全大補湯','正體類要',''),
    ('FM15','Restore the Spleen Decoction','歸脾湯','Gui Pi Tang','Tonify Qi+Blood','黃耆 12; 龍眼肉 9; 人參 9; 白朮 9; 當歸 6; 茯神 9; 遠志 6; 酸棗仁 9; 木香 3; 炙甘草 3; 生薑; 大棗','SY11; SY29','Tonify heart-spleen; nourish blood; calm shen','Yin deficiency with empty heat','—','濟生方',''),
    ('FM16','Heavenly Emperor Special Pill','天王補心丹','Tian Wang Bu Xin Dan','Calm Spirit','生地 12; 人參 6; 玄參 6; 丹參 6; 茯苓 6; 當歸 9; 五味子 6; 麥冬 9; 天冬 9; 柏子仁 9; 酸棗仁 9; 遠志 6; 桔梗 6','SY12','Nourish yin; clear empty heat; calm shen','Spleen-stomach weakness with damp','—','攝生秘剖',''),
    ('FM17','Four Gentlemen Decoction','四君子湯','Si Jun Zi Tang','Tonify Qi','人參 9; 白朮 9; 茯苓 9; 炙甘草 6','SY13','Tonify spleen qi','Yin deficiency','+陳皮+半夏 = 六君子湯; +木香+砂仁 = 香砂六君子','和劑局方',''),
    ('FM18','Six Gentlemen Decoction','六君子湯','Liu Jun Zi Tang','Tonify Qi','四君子 + 陳皮 6; 半夏 9','SY13 with phlegm','Tonify spleen qi; transform phlegm','Yin deficiency','—','醫學正傳',''),
    ('FM19','Tonify Middle Boost Qi','補中益氣湯','Bu Zhong Yi Qi Tang','Tonify Qi','黃耆 18; 人參 6; 白朮 9; 炙甘草 9; 當歸 3; 陳皮 6; 升麻 6; 柴胡 6','SY15','Lift middle qi; tonify','High fever; yin deficiency','—','脾胃論',''),
    ('FM20','Augment the Stomach Decoction','益胃湯','Yi Wei Tang','Tonify Yin','沙參 9; 麥冬 15; 生地 15; 玉竹 4.5; 冰糖 3','SY16','Nourish stomach yin','Cold-damp middle','—','溫病條辨',''),
    ('FM21','Preserve Harmony Pill','保和丸','Bao He Wan','Disperse Food','山楂 18; 神曲 6; 半夏 9; 茯苓 9; 陳皮 3; 連翹 3; 萊菔子 3','SY17','Disperse food; rectify qi; transform phlegm','Spleen deficiency','—','丹溪心法',''),
    ('FM22','Jade Windscreen Powder','玉屏風散','Yu Ping Feng San','Tonify Qi — Wei','黃耆 18; 白朮 9; 防風 6','SY18; recurrent colds','Tonify wei qi; secure exterior','Acute exterior pattern','—','丹溪心法',''),
    ('FM23','Glehnia Ophiopogon Decoction','沙參麥冬湯','Sha Shen Mai Dong Tang','Tonify Yin','北沙參 9; 麥冬 9; 玉竹 6; 桑葉 4.5; 天花粉 4.5; 扁豆 4.5; 甘草 3','SY19','Nourish lung-stomach yin','Cold-damp; spleen yang deficiency','—','溫病條辨',''),
    ('FM24','Three Unbinding Decoction','三拗湯','San Ao Tang','Disperse Lung','麻黃 9; 杏仁 9; 炙甘草 6','SY20','Disperse wind-cold; release lung','Wind-heat; sweating','—','和劑局方','CAUTION hypertension'),
    ('FM25','Stop Coughing Powder','止嗽散','Zhi Sou San','Disperse Lung','桔梗 6; 荊芥 6; 紫菀 6; 百部 6; 白前 6; 陳皮 3; 甘草 3','SY20 (chronic)','Disperse wind; descend lung qi; stop cough','Internal heat predominating','—','醫學心悟',''),
    ('FM26','Mulberry Chrysanthemum Drink','桑菊飲','Sang Ju Yin','Disperse Lung — Cool','桑葉 7.5; 菊花 3; 杏仁 6; 連翹 5; 桔梗 6; 蘆根 6; 薄荷 2.5; 甘草 2.5','SY21 (mild)','Disperse wind-heat; clear lung','Severe high fever','+石膏+知母 if heat strong','溫病條辨',''),
    ('FM27','Honeysuckle Forsythia Powder','銀翹散','Yin Qiao San','Disperse Lung — Cool','金銀花 9; 連翹 9; 桔梗 6; 薄荷 6; 牛蒡子 9; 淡豆豉 5; 荊芥穗 5; 蘆根 9; 竹葉 4; 甘草 5','SY21','Disperse wind-heat; clear toxin','Wind-cold pattern','—','溫病條辨',''),
    ('FM28','Mulberry Apricot Decoction','桑杏湯','Sang Xing Tang','Disperse Lung — Dryness','桑葉 3; 杏仁 4.5; 北沙參 6; 象貝 3; 香豉 3; 梔皮 3; 梨皮 3','Dryness Harming Lung','Disperse external dryness; nourish lung','Wet-damp pattern','—','溫病條辨',''),
    ('FM29','Kidney Qi Pill — Golden Cabinet','金匱腎氣丸','Jin Gui Shen Qi Wan','Tonify Yang','熟地 24; 山藥 12; 山茱萸 12; 澤瀉 9; 茯苓 9; 牡丹皮 9; 桂枝 3; 附子 3','SY22','Warm and tonify kidney yang','Yin deficiency with heat','—','金匱要略','附子 dose conservative'),
    ('FM30','Six-Ingredient Pill — Rehmannia','六味地黃丸','Liu Wei Di Huang Wan','Tonify Yin','熟地 24; 山藥 12; 山茱萸 12; 澤瀉 9; 茯苓 9; 牡丹皮 9','SY23','Nourish kidney yin','Spleen yang deficiency with damp','+知母+黃柏 = 知柏地黃; +枸杞+菊花 = 杞菊地黃','小兒藥證直訣',''),
    ('FM31','Anemarrhena-Phellodendron Pill','知柏地黃丸','Zhi Bai Di Huang Wan','Tonify Yin + Clear Heat','六味 + 知母 6; 黃柏 6','SY24','Nourish yin; clear empty heat','Spleen yang deficiency','—','醫宗金鑑',''),
    ('FM32','Lycium-Chrysanthemum Pill','杞菊地黃丸','Qi Ju Di Huang Wan','Tonify Yin — Liver-Kidney','六味 + 枸杞 9; 菊花 9','SY23 with eye sx','Nourish liver-kidney yin; brighten eyes','Wind-cold pattern','—','醫級',''),
    ('FM33','Two-Cured Decoction','二陳湯','Er Chen Tang','Transform Phlegm','半夏 9; 陳皮 9; 茯苓 9; 炙甘草 4.5; 生薑; 烏梅','SY26','Dry damp; transform phlegm; rectify qi; harmonise stomach','Yin deficiency; dry cough','+杏仁+蘇子 for cough','和劑局方',''),
    ('FM34','Calm the Stomach Powder','平胃散','Ping Wei San','Aromatic Damp','蒼朮 9; 厚朴 6; 陳皮 6; 炙甘草 3','SY26 (middle)','Dry damp; rectify qi; harmonise stomach','Yin deficiency','—','和劑局方',''),
    ('FM35','Three Kernels Decoction','三仁湯','San Ren Tang','Damp-Heat','杏仁 9; 白蔻仁 6; 薏苡仁 18; 半夏 9; 厚朴 6; 通草 6; 滑石 18; 竹葉 6','SY28','Disperse and transform damp; clear heat','Pure heat without damp','—','溫病條辨',''),
    ('FM36','Eight-Herb Powder for Rectification','八正散','Ba Zheng San','Damp-Heat — Lower','車前子 9; 瞿麥 9; 萹蓄 9; 滑石 12; 山梔子 6; 炙甘草 6; 木通 6; 大黃 6; 燈心','SY27','Clear heat; drain damp; promote urination','Pregnancy; spleen deficiency','—','和劑局方',''),
    ('FM37','Pinellia Magnolia Decoction','半夏厚朴湯','Ban Xia Hou Po Tang','Regulate Qi','半夏 12; 厚朴 9; 茯苓 12; 生薑 15; 紫蘇 6','Plum-pit qi (梅核氣)','Rectify qi; descend rebellion; transform phlegm','Yin deficiency; dry cough','—','金匱要略',''),
    ('FM38','Licorice-Wheat-Jujube Decoction','甘麥大棗湯','Gan Mai Da Zao Tang','Calm Spirit','甘草 9; 小麥 30; 大棗 10','Visceral agitation','Nourish heart; calm spirit; harmonise middle','—','—','金匱要略',''),
    ('FM39','Sour Jujube Decoction','酸棗仁湯','Suan Zao Ren Tang','Calm Spirit','酸棗仁 18; 茯苓 6; 知母 6; 川芎 3; 甘草 3','Liver yin def insomnia','Nourish blood; calm shen; clear empty heat','Phlegm-fire insomnia','—','金匱要略',''),
    ('FM40','Ephedra-Apricot-Gypsum-Licorice','麻杏石甘湯','Ma Xing Shi Gan Tang','Disperse Lung — Heat','麻黃 6; 杏仁 9; 石膏 18-24; 炙甘草 6','Lung heat (asthma/bronchitis)','Disperse lung; clear heat; relieve asthma','Wind-cold without heat','—','傷寒論',''),
    ('FM41','Minor Bluegreen Dragon Decoction','小青龍湯','Xiao Qing Long Tang','Release Exterior + Phlegm','麻黃 9; 桂枝 9; 細辛 3; 乾薑 9; 半夏 9; 五味子 6; 白芍 9; 炙甘草 6','Exterior cold + interior fluid','Release exterior; warm and transform fluid','Yin deficiency; heat','—','傷寒論','CAUTION elderly + cardiac'),
    ('FM42','Poria-Cinnamon-Atractylodes-Licorice','苓桂朮甘湯','Ling Gui Zhu Gan Tang','Phlegm-Damp','茯苓 12; 桂枝 9; 白朮 6; 炙甘草 6','Phlegm-fluid in middle','Warm yang; transform fluid; harmonise middle','Yin deficiency','—','傷寒論',''),
    ('FM43','Five Ling Powder','五苓散','Wu Ling San','Drain Damp','豬苓 9; 茯苓 9; 白朮 9; 澤瀉 15; 桂枝 6','Tai yang water retention','Promote urination; warm yang; transform qi','Yin def; damp-heat predominating','—','傷寒論',''),
    ('FM44','True Warrior Decoction','真武湯','Zhen Wu Tang','Warm Yang — Water','附子 9; 白朮 6; 茯苓 9; 白芍 9; 生薑 9','Spleen-kidney yang def + water','Warm yang; promote urination; harmonise blood','Yin deficiency','—','傷寒論','附子 toxicity caution'),
    ('FM45','Cinnamon Twig + Licorice Decoction','桂枝甘草湯','Gui Zhi Gan Cao Tang','Tonify Heart Yang','桂枝 12; 炙甘草 6','Heart yang def palpitations','Warm and tonify heart yang','Heat patterns','—','傷寒論',''),
    ('FM46','Aucklandia-Amomum Six Gentlemen','香砂六君子湯','Xiang Sha Liu Jun Zi Tang','Tonify Qi','六君子 + 木香 3; 砂仁 3','SY13 with qi stagnation','Tonify spleen qi; rectify qi; harmonise stomach','Yin deficiency','—','古今名醫方論',''),
    ('FM47','Modified Major Bupleurum','大柴胡湯','Da Chai Hu Tang','Harmonise + Drain','柴胡 12; 黃芩 9; 半夏 9; 白芍 9; 枳實 9; 大黃 6; 生薑 9; 大棗','Shao yang + yang ming combined','Harmonise shao yang; drain interior','Pure exterior or pure deficiency','—','傷寒論',''),
    ('FM48','Pinellia Heart-Decoction','半夏瀉心湯','Ban Xia Xie Xin Tang','Harmonise — Stomach','半夏 12; 黃芩 9; 黃連 3; 乾薑 9; 人參 9; 大棗 4; 炙甘草 6','Stomach lump epigastric','Harmonise cold and heat; descend rebellion','Pure cold or pure heat','—','傷寒論',''),
    ('FM49','All-Inclusive Great Tonifying','十全大補湯','Shi Quan Da Bu Tang','Tonify Qi+Blood','八珍 + 黃耆 12; 肉桂 3','SY25 (severe)','Greatly tonify qi-blood-yang','Yin deficiency with heat','—','和劑局方',''),
    ('FM50','Mume Pill','烏梅丸','Wu Mei Wan','Cold-Heat','烏梅; 細辛; 乾薑; 黃連; 當歸; 附子; 蜀椒; 桂枝; 人參; 黃柏 (classical doses)','Roundworm; cold-heat jueyin','Calm worms; harmonise cold-heat','Pure cold or pure heat','—','傷寒論','VERIFY classical doses; pregnancy caution'),
]
write_rows(ws, formulas)

# Herbs
ws = wb.create_sheet('Herbs')
setup_sheet(ws,
    ['ID','EN Name','ZH Name','PY','Latin','Category','Taste (味)','Temp (性)','Meridian (歸經)','Function (功能)','Daily Dose (g)','Cautions'],
    [10, 30, 12, 22, 32, 24, 14, 12, 22, 50, 12, 35])

herbs = [
    ('HB01','Ginseng (Asian)','人參','Ren Shen','Panax ginseng','Tonify Qi','sweet, slightly bitter','slightly warm','LU SP HT','Greatly tonifies original qi; generates fluids; calms shen','3-9','Heat patterns; high BP'),
    ('HB02','Codonopsis','黨參','Dang Shen','Codonopsis pilosula','Tonify Qi','sweet','neutral','LU SP','Tonifies middle qi; gentler substitute for ginseng','9-30','—'),
    ('HB03','Pseudostellaria','太子參','Tai Zi Shen','Pseudostellaria heterophylla','Tonify Qi','sweet, slightly bitter','neutral','LU SP','Tonifies qi; nourishes yin; mild for children','9-30','—'),
    ('HB04','Astragalus','黃耆','Huang Qi','Astragalus membranaceus','Tonify Qi','sweet','slightly warm','LU SP','Tonifies wei qi; raises yang; promotes urination; promotes healing','9-30','Acute exterior pattern; heat'),
    ('HB05','Atractylodes (white)','白朮','Bai Zhu','Atractylodes macrocephala','Tonify Qi','sweet, bitter','warm','SP ST','Tonifies spleen; dries damp; secures wei','6-15','Yin deficiency; dryness'),
    ('HB06','Chinese yam','山藥','Shan Yao','Dioscorea opposita','Tonify Qi','sweet','neutral','LU SP KD','Tonifies spleen-lung-kidney; nourishes yin','9-30','Excess pattern with constipation'),
    ('HB07','Liquorice (raw)','甘草','Gan Cao','Glycyrrhiza uralensis','Tonify Qi','sweet','neutral','12 channels','Harmonises herbs; tonifies middle; resolves toxin','3-9','Long-term high dose oedema/HTN'),
    ('HB08','Liquorice (honey-fried)','炙甘草','Zhi Gan Cao','—','Tonify Qi','sweet','warm','12 channels','Tonifies middle; warms; harmonises','3-9','Same as above'),
    ('HB09','Jujube','大棗','Da Zao','Ziziphus jujuba','Tonify Qi','sweet','warm','SP ST','Tonifies middle; nourishes blood; calms shen','3-12','Damp-phlegm; food stagnation'),
    ('HB10','Tang-kuei','當歸','Dang Gui','Angelica sinensis','Tonify Blood','sweet, acrid','warm','LV HT SP','Tonifies blood; invigorates blood; regulates menses; moistens intestines','6-15','Damp-phlegm with diarrhoea; pregnancy caution'),
    ('HB11','Cooked rehmannia','熟地黃','Shu Di Huang','Rehmannia glutinosa (prep)','Tonify Blood','sweet','slightly warm','LV KD','Nourishes blood and yin; tonifies essence','9-30','Damp-phlegm; spleen yang def'),
    ('HB12','White peony','白芍','Bai Shao','Paeonia lactiflora','Tonify Blood','sour, bitter','slightly cold','LV SP','Nourishes blood; preserves yin; soothes liver; relieves pain','6-15','Cold-damp diarrhoea'),
    ('HB13','Donkey-hide gelatin','阿膠','E Jiao','Asini Corii Colla','Tonify Blood','sweet','neutral','LU LV KD','Nourishes blood; nourishes yin; stops bleeding','3-9','Damp-phlegm; halal NON-COMPLIANT'),
    ('HB14','Longan flesh','龍眼肉','Long Yan Rou','Dimocarpus longan','Tonify Blood','sweet','warm','HT SP','Nourishes blood; tonifies heart-spleen; calms shen','9-15','Damp-phlegm; food stagnation'),
    ('HB15','Glehnia (north)','北沙參','Bei Sha Shen','Glehnia littoralis','Tonify Yin','sweet, slightly bitter','slightly cold','LU ST','Nourishes lung-stomach yin; clears empty heat','5-12','Wind-cold cough'),
    ('HB16','Ophiopogon','麥冬','Mai Dong','Ophiopogon japonicus','Tonify Yin','sweet, slightly bitter','slightly cold','LU HT ST','Moistens lung; nourishes yin; calms shen','6-12','Cold-damp; spleen yang def'),
    ('HB17','Asparagus root','天冬','Tian Dong','Asparagus cochinchinensis','Tonify Yin','sweet, bitter','cold','LU KD','Nourishes lung-kidney yin; clears empty heat','6-12','Spleen yang def with diarrhoea'),
    ('HB18','Solomon seal','玉竹','Yu Zhu','Polygonatum odoratum','Tonify Yin','sweet','slightly cold','LU ST','Nourishes lung-stomach yin; generates fluids','6-12','Damp-phlegm'),
    ('HB19','Lycium fruit','枸杞子','Gou Qi Zi','Lycium barbarum','Tonify Yin','sweet','neutral','LV KD','Nourishes liver-kidney yin; brightens eyes','6-15','Damp-phlegm with diarrhoea'),
    ('HB20','Privet fruit','女貞子','Nu Zhen Zi','Ligustrum lucidum','Tonify Yin','sweet, bitter','cool','LV KD','Nourishes liver-kidney yin','6-15','Spleen yang def with diarrhoea'),
    ('HB21','Deer antler velvet','鹿茸','Lu Rong','Cervus nippon','Tonify Yang','sweet, salty','warm','LV KD','Tonifies kidney yang; strengthens essence; warms ming men','1-3','Yin def heat; pregnancy; halal NON-COMPLIANT'),
    ('HB22','Epimedium','淫羊藿','Yin Yang Huo','Epimedium brevicornu','Tonify Yang','acrid, sweet','warm','LV KD','Tonifies kidney yang; expels wind-damp','3-9','Yin deficiency heat'),
    ('HB23','Morinda root','巴戟天','Ba Ji Tian','Morinda officinalis','Tonify Yang','acrid, sweet','slightly warm','LV KD','Tonifies kidney yang; strengthens sinew-bone','3-9','Yin deficiency'),
    ('HB24','Eucommia','杜仲','Du Zhong','Eucommia ulmoides','Tonify Yang','sweet, slightly acrid','warm','LV KD','Tonifies kidney; strengthens lower back; calms fetus','9-15','Yin deficiency heat'),
    ('HB25','Ephedra','麻黃','Ma Huang','Ephedra sinica','Release Exterior — Warm','acrid, slightly bitter','warm','LU UB','Disperses wind-cold; promotes sweat; relieves asthma','1.5-9','Sweating; HTN; cardiac; insomnia; sport doping'),
    ('HB26','Cinnamon twig','桂枝','Gui Zhi','Cinnamomum cassia','Release Exterior — Warm','acrid, sweet','warm','HT LU UB','Releases exterior; warms channels; warms yang','3-9','Heat patterns; pregnancy bleeding'),
    ('HB27','Perilla leaf','紫蘇葉','Zi Su Ye','Perilla frutescens','Release Exterior — Warm','acrid','warm','LU SP','Releases exterior; rectifies qi; calms fetus; resolves seafood toxin','3-9','Long cooking destroys volatile oils'),
    ('HB28','Fresh ginger','生薑','Sheng Jiang','Zingiber officinale','Release Exterior — Warm','acrid','slightly warm','LU SP ST','Releases exterior; warms middle; resolves toxin','3-9','Heat patterns; severe yin def'),
    ('HB29','Saposhnikovia','防風','Fang Feng','Saposhnikovia divaricata','Release Exterior — Warm','acrid, sweet','slightly warm','UB LV SP','Releases exterior; expels wind; relieves spasm','3-9','Yin def with wind from blood def'),
    ('HB30','Mulberry leaf','桑葉','Sang Ye','Morus alba','Release Exterior — Cool','sweet, bitter','cold','LU LV','Disperses wind-heat; clears liver; brightens eyes','5-9','Wind-cold cough'),
    ('HB31','Chrysanthemum','菊花','Ju Hua','Chrysanthemum morifolium','Release Exterior — Cool','sweet, bitter, acrid','slightly cold','LU LV','Disperses wind-heat; clears liver; resolves toxin','5-9','Cold-damp pattern'),
    ('HB32','Mint','薄荷','Bo He','Mentha haplocalyx','Release Exterior — Cool','acrid','cool','LU LV','Disperses wind-heat; clears head-eye; courses liver qi','3-6','Late addition; do not boil long'),
    ('HB33','Burdock seed','牛蒡子','Niu Bang Zi','Arctium lappa','Release Exterior — Cool','acrid, bitter','cold','LU ST','Disperses wind-heat; resolves toxin; benefits throat','6-12','Spleen def with diarrhoea'),
    ('HB34','Bupleurum','柴胡','Chai Hu','Bupleurum chinense','Release Exterior — Cool','bitter, acrid','slightly cold','LV GB PC TB','Releases shao yang; soothes liver; raises yang','3-9','Liver fire; yin def'),
    ('HB35','Kudzu root','葛根','Ge Gen','Pueraria lobata','Release Exterior — Cool','sweet, acrid','cool','SP ST','Releases muscles; generates fluids; raises yang','9-15','Yang def with diarrhoea'),
    ('HB36','Gypsum','石膏','Shi Gao','Gypsum (mineral)','Clear Heat — Qi level','sweet, acrid','very cold','LU ST','Clears heat from qi level; relieves thirst','15-60','Cold patterns; yang deficiency'),
    ('HB37','Anemarrhena','知母','Zhi Mu','Anemarrhena asphodeloides','Clear Heat — Yin','bitter, sweet','cold','LU ST KD','Clears heat; nourishes yin; relieves bone-steaming','6-12','Spleen yang def with diarrhoea'),
    ('HB38','Coptis','黃連','Huang Lian','Coptis chinensis','Clear Heat — Damp','bitter','cold','HT LV ST LI','Clears damp-heat; drains fire; resolves toxin','2-5','Long high doses spleen damage'),
    ('HB39','Scutellaria','黃芩','Huang Qin','Scutellaria baicalensis','Clear Heat — Damp','bitter','cold','LU GB ST LI','Clears damp-heat; drains fire; resolves toxin; calms fetus','3-9','Spleen-stomach cold deficiency'),
    ('HB40','Phellodendron','黃柏','Huang Bai','Phellodendron amurense','Clear Heat — Damp','bitter','cold','KD UB LI','Clears damp-heat in lower; drains kidney fire','3-12','Spleen-stomach cold def'),
    ('HB41','Forsythia','連翹','Lian Qiao','Forsythia suspensa','Clear Heat — Toxin','bitter','slightly cold','LU HT GB','Clears heat; resolves toxin; reduces nodules','6-15','Spleen def with diarrhoea'),
    ('HB42','Honeysuckle','金銀花','Jin Yin Hua','Lonicera japonica','Clear Heat — Toxin','sweet','cold','LU HT ST','Clears heat; resolves toxin; disperses wind-heat','6-15','Spleen-stomach cold def'),
    ('HB43','Gardenia fruit','梔子','Zhi Zi','Gardenia jasminoides','Clear Heat — Fire','bitter','cold','HT LU ST SJ','Drains fire; clears damp-heat; cools blood','3-9','Spleen def with diarrhoea'),
    ('HB44','Rhubarb','大黃','Da Huang','Rheum palmatum','Drain Downward','bitter','cold','SP ST LI LV','Drains heat; purges accumulation; cools blood; invigorates blood','3-12','Pregnancy; menstruation; weak elderly'),
    ('HB45','Mirabilite','芒硝','Mang Xiao','Mirabilite (mineral)','Drain Downward','salty, bitter','cold','LI ST','Drains heat; softens hardness','6-9','Pregnancy; weak constitution'),
    ('HB46','Poria','茯苓','Fu Ling','Poria cocos','Drain Damp','sweet, bland','neutral','HT LU SP KD','Promotes urination; tonifies spleen; calms shen','9-15','Yin def without damp'),
    ('HB47','Polyporus','豬苓','Zhu Ling','Polyporus umbellatus','Drain Damp','sweet, bland','neutral','KD UB','Promotes urination; drains damp','6-12','—'),
    ('HB48','Alisma','澤瀉','Ze Xie','Alisma orientale','Drain Damp','sweet, bland','cold','KD UB','Promotes urination; drains damp; clears empty heat','5-10','—'),
    ('HB49','Coix seed','薏苡仁','Yi Yi Ren','Coix lacryma-jobi','Drain Damp','sweet, bland','cool','SP ST LU','Promotes urination; tonifies spleen; clears heat; expels pus','9-30','Pregnancy (causes uterine contraction)'),
    ('HB50','Plantago seed','車前子','Che Qian Zi','Plantago asiatica','Drain Damp','sweet','cold','LV KD UB SI','Promotes urination; clears damp-heat; cools liver','9-15','Wrap in cloth bag for cooking'),
    ('HB51','Cangzhu','蒼朮','Cang Zhu','Atractylodes lancea','Aromatic Damp','acrid, bitter','warm','SP ST','Dries damp; strengthens spleen; dispels wind-cold','3-9','Yin def; sweating'),
    ('HB52','Magnolia bark','厚朴','Hou Po','Magnolia officinalis','Aromatic Damp','bitter, acrid','warm','SP ST LU LI','Dries damp; rectifies qi; descends rebellious qi','3-9','Pregnancy'),
    ('HB53','Patchouli','藿香','Huo Xiang','Pogostemon cablin','Aromatic Damp','acrid','slightly warm','SP ST LU','Aromatically transforms damp; harmonises stomach; releases exterior','5-10','—'),
    ('HB54','Amomum fruit','砂仁','Sha Ren','Amomum villosum','Aromatic Damp','acrid','warm','SP ST KD','Transforms damp; rectifies qi; harmonises stomach; calms fetus','3-6','Yin def; high temp short cook'),
    ('HB55','Tangerine peel','陳皮','Chen Pi','Citrus reticulata (aged)','Regulate Qi','acrid, bitter','warm','LU SP','Rectifies qi; harmonises middle; dries damp','3-9','Yin def; dry cough'),
    ('HB56','Cyperus','香附','Xiang Fu','Cyperus rotundus','Regulate Qi','acrid, slightly bitter','neutral','LV TB','Soothes liver; rectifies qi; regulates menses','6-12','Yin def with internal heat'),
    ('HB57','Aucklandia','木香','Mu Xiang','Saussurea lappa','Regulate Qi','acrid, bitter','warm','SP ST LI GB SJ','Rectifies qi; harmonises stomach','3-9','Yin def with heat'),
    ('HB58','Bitter orange (immature)','枳實','Zhi Shi','Citrus aurantium (immature)','Regulate Qi','bitter, acrid, sour','slightly cold','SP ST LI','Breaks qi stagnation; transforms phlegm; reduces accumulation','3-9','Pregnancy; weak'),
    ('HB59','Bitter orange (mature)','枳殼','Zhi Ke','Citrus aurantium (mature)','Regulate Qi','bitter, acrid, sour','slightly cold','SP ST LI','Rectifies qi; gentler than 枳實','3-9','Pregnancy'),
    ('HB60','Sichuan lovage','川芎','Chuan Xiong','Ligusticum chuanxiong','Invigorate Blood','acrid','warm','LV GB PC','Invigorates blood; rectifies qi; expels wind; relieves pain','3-9','Yin def headache; menorrhagia'),
    ('HB61','Salvia (Chinese)','丹參','Dan Shen','Salvia miltiorrhiza','Invigorate Blood','bitter','slightly cold','HT PC','Invigorates blood; cools blood; calms shen','9-15','Pregnancy; bleeding; warfarin interaction'),
    ('HB62','Peach kernel','桃仁','Tao Ren','Prunus persica','Invigorate Blood','bitter, sweet','neutral','HT LV LU LI','Invigorates blood; breaks stasis; moistens intestines','5-9','Pregnancy; diarrhoea'),
    ('HB63','Safflower','紅花','Hong Hua','Carthamus tinctorius','Invigorate Blood','acrid','warm','HT LV','Invigorates blood; resolves stasis; promotes menses','3-9','Pregnancy; bleeding'),
    ('HB64','Motherwort','益母草','Yi Mu Cao','Leonurus japonicus','Invigorate Blood','acrid, bitter','slightly cold','LV PC UB','Invigorates blood; regulates menses; promotes urination','9-30','Pregnancy'),
    ('HB65','Notoginseng','三七','San Qi','Panax notoginseng','Stop Bleeding','sweet, slightly bitter','warm','LV ST','Stops bleeding; transforms stasis; reduces swelling','3-9 (1-3 powder)','Pregnancy'),
    ('HB66','Schizandra','五味子','Wu Wei Zi','Schisandra chinensis','Astringent','sour','warm','LU HT KD','Astringes lung; tonifies kidney; generates fluids; calms shen','2-6','Acute exterior; heat'),
    ('HB67','Cornus','山茱萸','Shan Zhu Yu','Cornus officinalis','Astringent','sour, astringent','slightly warm','LV KD','Tonifies liver-kidney; secures jing; stops sweat','6-12','Damp-heat; difficult urination'),
    ('HB68','Mume fruit','烏梅','Wu Mei','Prunus mume','Astringent','sour','neutral','LV SP LU LI','Astringes lung; quiets parasites; generates fluids','3-9','Acute exterior'),
    ('HB69','Sour jujube seed','酸棗仁','Suan Zao Ren','Ziziphus spinosa','Calm Spirit','sweet, sour','neutral','HT LV GB','Nourishes heart; calms shen; stops sweat','9-15','Liver fire; phlegm-fire insomnia'),
    ('HB70','Polygala','遠志','Yuan Zhi','Polygala tenuifolia','Calm Spirit','acrid, bitter','slightly warm','HT KD LU','Calms shen; opens orifices; transforms phlegm','3-9','Gastric ulcer; gastritis'),
    ('HB71','Dragon bone','龍骨','Long Gu','Os draconis (fossil)','Calm Spirit — Heavy','sweet, astringent','neutral','HT LV KD','Calms shen; subdues yang; astringes','15-30','Damp-heat; halal NON-COMPLIANT'),
    ('HB72','Oyster shell','牡蠣','Mu Li','Ostrea','Calm Spirit — Heavy','salty','slightly cold','LV KD','Calms shen; subdues yang; softens hardness; astringes','15-30','Spleen def; halal NON-COMPLIANT'),
    ('HB73','Pinellia','半夏','Ban Xia','Pinellia ternata (prep)','Transform Phlegm','acrid','warm','SP ST LU','Dries damp; transforms phlegm; descends rebellion; reduces nodules','3-9','Pregnancy; bleeding; with 烏頭 (incompatible)'),
    ('HB74','Platycodon','桔梗','Jie Geng','Platycodon grandiflorus','Transform Phlegm','bitter, acrid','neutral','LU','Disperses lung qi; expels phlegm; benefits throat','3-9','—'),
    ('HB75','Apricot kernel','杏仁','Xing Ren','Prunus armeniaca','Transform Phlegm','bitter','slightly warm','LU LI','Stops cough; descends lung qi; moistens intestines','3-9','Toxic raw; remove tip'),
    ('HB76','Fritillary (Zhejiang)','浙貝母','Zhe Bei Mu','Fritillaria thunbergii','Transform Phlegm — Heat','bitter','cold','LU HT','Clears heat; transforms phlegm; reduces nodules','3-9','—'),
    ('HB77','Fritillary (Sichuan)','川貝母','Chuan Bei Mu','Fritillaria cirrhosa','Transform Phlegm — Heat','bitter, sweet','slightly cold','LU HT','Clears heat; moistens lung; transforms phlegm','3-9','Spleen-stomach cold-damp'),
    ('HB78','Aconite (prepared)','附子','Fu Zi','Aconitum carmichaelii (prep)','Warm Interior','acrid, sweet','very hot','HT KD SP','Restores devastated yang; warms ming men; dispels cold-damp','3-15','Pregnancy; toxicity if undercooked; 先煎 30-60 min'),
    ('HB79','Dried ginger','乾薑','Gan Jiang','Zingiber officinale (dried)','Warm Interior','acrid','hot','SP ST KD HT LU','Warms middle; rescues yang; warms lung; transforms fluid','3-9','Yin def with heat; pregnancy caution'),
    ('HB80','Cinnamon bark','肉桂','Rou Gui','Cinnamomum cassia (bark)','Warm Interior','acrid, sweet','very hot','KD SP HT LV','Warms ming men; restores yang; disperses cold; promotes warm circulation','1-5','Pregnancy; bleeding; hot patterns'),
]
write_rows(ws, herbs)

# Symptom dictionary
ws = wb.create_sheet('Symptoms')
setup_sheet(ws,
    ['ID','Category','Standard Term EN','Standard Term ZH','Synonyms / Patient-spoken Variants','Maps to Syndrome IDs (criteria match)'],
    [10, 22, 30, 18, 60, 30])

symptoms = [
    ('SX01','Body temperature','Fever','發熱','warm; hot; pyrexia; raised temp','SY01; SY02; SY03; SY04; SY21'),
    ('SX02','Body temperature','Aversion to cold','惡寒','cold-feeling; chills; cant warm up','SY01; SY02; SY05; SY14'),
    ('SX03','Body temperature','Aversion to wind','惡風','wind-sensitive; draft-sensitive','SY02; SY18'),
    ('SX04','Sweating','Spontaneous sweating','自汗','sweats easily; sweats at rest','SY18; SY13; SY15'),
    ('SX05','Sweating','Night sweats','盜汗','sweats during sleep','SY12; SY19; SY23; SY24'),
    ('SX06','Sweating','Absence of sweat','無汗','dry skin; doesnt sweat','SY01'),
    ('SX07','Pain','Headache','頭痛','head pain; migraine','SY01; SY09; SY21'),
    ('SX08','Pain','Hypochondriac distension','脅脹','side-rib pressure; chest-side fullness','SY07; SY08; SY30'),
    ('SX09','Pain','Lower back pain','腰痛','lumbar ache; lower back soreness','SY22; SY23'),
    ('SX10','Mental','Irritability','煩躁','easy anger; easily upset','SY07; SY08; SY09'),
    ('SX11','Mental','Insomnia','失眠','cant sleep; trouble sleeping','SY11; SY12; SY24; SY29'),
    ('SX12','Mental','Vivid dreams','多夢','many dreams; nightmares','SY11; SY12'),
    ('SX13','Mental','Palpitations','心悸','heart racing; awareness of heartbeat','SY11; SY12; SY29'),
    ('SX14','Digestive','Abdominal bloating','腹脹','belly fullness; distension after eating','SY13; SY15; SY17; SY30'),
    ('SX15','Digestive','Loose stool','便溏','soft stool; diarrhoea-like','SY13; SY14; SY29; SY30'),
    ('SX16','Digestive','Constipation','便秘','hard stool; difficulty passing','SY04; SY16; SY24'),
    ('SX17','Digestive','Poor appetite','納差','no appetite; eat less','SY13; SY15; SY29'),
    ('SX18','Digestive','Hunger without appetite','嘈雜','hungry but cant eat','SY16'),
    ('SX19','Digestive','Sour belching','噯腐','sour reflux; food-smelling burp','SY17'),
    ('SX20','Respiratory','Cough — productive','咳嗽有痰','wet cough; sputum present','SY20; SY21; SY26'),
    ('SX21','Respiratory','Cough — dry','乾咳','dry cough; no sputum','SY19'),
    ('SX22','Respiratory','Asthma / SOB','喘','wheezing; shortness of breath','SY18; SY26'),
    ('SX23','Respiratory','Sore throat','咽痛','throat pain; pharyngitis-like','SY21; SY24'),
    ('SX24','Genitourinary','Frequent clear urination','小便清長','urinates a lot; clear pale urine','SY22'),
    ('SX25','Genitourinary','Yellow turbid urine','小便黃濁','dark urine; cloudy urine','SY27'),
    ('SX26','Genitourinary','Impotence','陽痿','erectile difficulty','SY22'),
    ('SX27','Womens health','Scant menses','月經量少','light period; reduced flow','SY10; SY11; SY25'),
    ('SX28','Womens health','PMS / breast distension','經前乳脹','pre-period mood; breast pressure','SY07; SY30'),
    ('SX29','General','Fatigue','疲勞','tired; lassitude; no energy','SY13; SY15; SY18; SY25; SY29'),
    ('SX30','General','Cold extremities','四肢厥冷','cold hands and feet','SY05; SY06; SY14; SY22'),
    ('SX31','General','Heavy body','身重','sluggish; heavy limbs','SY26; SY28'),
    ('SX32','General','Bitter mouth','口苦','bitter taste','SY03; SY08'),
    ('SX33','General','Dry mouth','口乾','dry mouth; thirsty','SY16; SY19; SY23; SY24'),
]
write_rows(ws, symptoms)

# Red flags
ws = wb.create_sheet('RedFlags')
setup_sheet(ws,
    ['ID','Type','Condition / Drug','Avoid Herbs / Formulas','Reason','Severity','Notes for Malaysian context'],
    [10, 22, 32, 50, 50, 16, 35])

redflags = [
    ('RF01','Pregnancy','Pregnancy — any trimester','大黃; 芒硝; 桃仁; 紅花; 益母草; 三七; 麝香; 蟾酥; 半夏 (CAUTION); 附子 (CAUTION)','Strong blood-movers and toxic herbs can induce miscarriage','Absolute',''),
    ('RF02','Pregnancy','Pregnancy — first trimester especially','薏苡仁','Causes uterine contraction','Relative','Used safely in late pregnancy by some doctors; verify'),
    ('RF03','Pregnancy','Pregnancy','大承氣湯; 桃紅四物湯','Strong purgatives + blood-movers','Absolute',''),
    ('RF04','Drug interaction','Warfarin / DOACs','丹參; 當歸; 三七; 紅花; 川芎','Increased bleeding risk via platelet/coagulation effects','Relative — monitor INR if used',''),
    ('RF05','Drug interaction','Digoxin','甘草 (long-term high dose)','Hypokalaemia from licorice → digoxin toxicity','Relative — limit dose and duration',''),
    ('RF06','Drug interaction','SSRIs / MAOIs','麻黃','Sympathomimetic interaction; hypertensive crisis','Absolute',''),
    ('RF07','Drug interaction','Antihypertensives','麻黃','Antagonises BP control','Absolute',''),
    ('RF08','Cardiac','Hypertension','麻黃; 細辛 (high dose); 甘草 (long-term high dose)','Raises BP / fluid retention','Relative — substitute or low-dose only',''),
    ('RF09','Cardiac','Cardiac arrhythmia','附子 (uncooked); 烏頭','Cardio-active alkaloids; arrhythmia risk','Absolute for raw; relative for prepared',''),
    ('RF10','Halal','Patient prefers halal-only','鹿茸; 龍骨; 牡蠣; 阿膠; 蛇膽; 蟾酥; 海馬; 蜈蚣; 全蠍','Animal-derived (deer, fossilised bones, oyster, donkey, snake, toad, seahorse, centipede, scorpion)','Patient preference','Substitute with plant-based equivalents; consult patient'),
    ('RF11','Halal','Patient prefers halal-only','酒製 (alcohol-processed) herbs e.g. 酒當歸','Alcohol-prepared variants violate halal','Patient preference','Use water-prepared variants; mention to dispensing pharmacy'),
    ('RF12','Paediatric','Children < 12','鹿茸; 麻黃 (high dose); 附子 (CAUTION)','Strong tonics may cause precocious development; ephedra cardiac risk','Relative','Reduce doses; prefer mild formulas'),
    ('RF13','Renal','Chronic kidney disease','關木通; 馬兜鈴 family; 細辛 (high)','Aristolochic acid → nephrotoxic; banned in many countries','Absolute','Substitute 木通 with 川木通'),
    ('RF14','Hepatic','Liver disease','薄荷油 high dose; 何首烏 (raw)','Hepatotoxicity reports','Relative','Use prepared 製何首烏 only; monitor LFTs'),
    ('RF15','GI','Active GI bleed','大黃; 桃仁; 紅花','Worsens bleeding','Absolute',''),
    ('RF16','Allergy','Known herb allergy','As reported','Cross-reactivity possible within plant families','Absolute','Document; avoid related herbs'),
]
write_rows(ws, redflags)

# Reorder so README is first tab
sheets = ['README','Syndromes','Formulas','Herbs','Symptoms','RedFlags']
wb._sheets = [wb[s] for s in sheets]

wb.save(OUT)
print('Saved', OUT)
print('Counts: syndromes=' + str(len(syndromes)) + ', formulas=' + str(len(formulas)) + ', herbs=' + str(len(herbs)) + ', symptoms=' + str(len(symptoms)) + ', redflags=' + str(len(redflags)))
